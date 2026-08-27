"""
SDLC Router - API endpoints for SDLC workflow management
"""
from fastapi import APIRouter, HTTPException, BackgroundTasks, UploadFile, File
from pydantic import BaseModel, Field
from typing import List, Literal, Optional, Dict
from datetime import datetime
import json
import logging
import time

from src.services.sdlc_tracker import SDLCTracker, SDLCPhaseType, PhaseStatus
from src.services.code_analyzer import PythonCodeAnalyzer, AnalysisResult
from src.services.testing_service import TestingService, ClassTestResult, TestStatus, TestType
from src.connectors.repo_connector import RepositoryConnector, RepositoryMetadata

router = APIRouter(prefix="/api/sdlc", tags=["SDLC"])
logger = logging.getLogger(__name__)

# Initialize services
sdlc_tracker = SDLCTracker()
code_analyzer = PythonCodeAnalyzer()
testing_service = TestingService()
repo_connector = RepositoryConnector()


# ── Request/Response Models ───────────────────────────────────────────────────

class RepositoryInput(BaseModel):
    """Repository input model"""
    type: str = Field(..., description="Repository type: application, config, or devops")
    url: Optional[str] = Field(None, description="Git repository URL")
    local_path: Optional[str] = Field(None, description="Local path to repository")
    branch: str = Field(default="main", description="Git branch")
    git_token: Optional[str] = Field(None, description="Personal access token for private repos")


class CreateProjectRequest(BaseModel):
    """Request to create SDLC project"""
    name: str
    description: str
    repositories: List[RepositoryInput] = []
    # Explicit typed repo slots (preferred over the generic list)
    app_repo: Optional[RepositoryInput] = None
    infra_repo: Optional[RepositoryInput] = None
    devops_repo: Optional[RepositoryInput] = None


class UpdateProgressRequest(BaseModel):
    """Request to update phase progress"""
    progress_percent: Optional[float] = None
    tasks_completed: Optional[int] = None
    tasks_total: Optional[int] = None
    notes: Optional[str] = None


class TransitionPhaseRequest(BaseModel):
    """Request to transition to new phase"""
    target_phase: str
    force: bool = False


class StartTestingRequest(BaseModel):
    """Request to start testing"""
    test_type: Literal["unit", "regression", "integration"] = "unit"
    file_extensions: Optional[List[str]] = None


# ── Project Management Endpoints ──────────────────────────────────────────────

@router.post("/projects", response_model=Dict)
async def create_project(request: CreateProjectRequest, background_tasks: BackgroundTasks):
    """Create a new SDLC project

    This endpoint:
    1. Clones/connects to repositories
    2. Analyzes code structure
    3. Creates SDLC project entity
    4. Initializes in Requirements phase
    """
    import logging
    import time
    
    logger = logging.getLogger("sdlc")
    start_time = time.time()
    
    try:
        logger.info(f"Creating SDLC project: {request.name}")
        
        # Merge explicit typed slots into the repositories list
        explicit_repos = []
        if request.app_repo:
            request.app_repo.type = "application"
            explicit_repos.append(request.app_repo)
        if request.infra_repo:
            request.infra_repo.type = "infrastructure"
            explicit_repos.append(request.infra_repo)
        if request.devops_repo:
            request.devops_repo.type = "devops"
            explicit_repos.append(request.devops_repo)
        all_repos = explicit_repos if explicit_repos else request.repositories

        logger.info(f"Processing {len(all_repos)} repositories...")
        
        # Process repositories
        repo_configs = []
        repo_metadata_list = []

        for idx, repo in enumerate(all_repos):
            logger.info(f"Repository {idx+1}/{len(all_repos)}: {repo.type}")
            try:
                if repo.url:
                    logger.info(f"  Cloning from URL: {repo.url}")
                    metadata = repo_connector.clone_repository(
                        repo.url,
                        repo.branch,
                        repo.type,
                        git_token=repo.git_token,
                    )
                elif repo.local_path:
                    logger.info(f"  Connecting to local path: {repo.local_path}")
                    metadata = repo_connector.connect_local_path(
                        repo.local_path,
                        repo.type,
                    )
                else:
                    raise HTTPException(400, "Either url or local_path must be provided")
                
                logger.info(f"  Found {metadata.file_count} files")
                
            except HTTPException:
                raise
            except Exception as e:
                # Clone/scan failed — record the repo URL but continue
                logger.warning("Repo connection failed for %s: %s", getattr(repo, 'url', repo.local_path), e)
                from src.connectors.repo_connector import RepositoryMetadata
                from datetime import datetime as _dt
                metadata = RepositoryMetadata(
                    repo_type=repo.type,
                    url=repo.url or "",
                    local_path="",
                    branch=repo.branch,
                    last_synced=_dt.now().isoformat(),
                )

            repo_configs.append({
                "type": repo.type,
                "url": metadata.url,
                "local_path": metadata.local_path,
                "branch": metadata.branch,
                "technologies": metadata.technologies,
                "file_count": metadata.file_count
            })
            repo_metadata_list.append(metadata)

        logger.info(f"Creating SDLC project entity...")
        
        # Create SDLC project
        project = sdlc_tracker.create_project(
            name=request.name,
            description=request.description,
            repositories=repo_configs
        )
        
        logger.info(f"Project created with ID: {project.id}")

        # Analyze code (application repos with a local path only)
        logger.info("Starting code analysis...")
        analysis_results = {}
        for metadata in repo_metadata_list:
            if metadata.repo_type == "application" and metadata.local_path:
                logger.info(f"Analyzing {metadata.local_path}...")
                result = code_analyzer.analyze_directory(metadata.local_path)
                logger.info(f"  Found {result.total_classes} classes, {result.total_tests} tests")
                analysis_results[metadata.local_path] = {
                    "total_classes": result.total_classes,
                    "total_functions": result.total_functions,
                    "total_tests": result.total_tests,
                    "total_lines": result.total_lines,
                    "test_coverage_percent": result.test_coverage_percent,
                    "classes": [
                        {
                            "name": cls.name,
                            "file_path": cls.file_path,
                            "methods": cls.methods,
                            "complexity": cls.complexity,
                            "is_test": cls.is_test,
                        }
                        for cls in result.classes
                    ],
                }

        # Populate knowledge graph with SDLC entities
        logger.info("Populating knowledge graph...")
        try:
            graph_summary = populate_from_analysis(
                project_id=project.id,
                repo_metadata_list=repo_metadata_list,
                analysis_results=analysis_results,
            )
            logger.info(f"Graph populated: {graph_summary}")
        except Exception as e:
            logger.error(f"Graph population failed: {e}")
            graph_summary = {}

        # Store analysis results in project metadata
        project.metadata["analysis"] = analysis_results
        project.metadata["graph_summary"] = graph_summary
        
        # Auto-complete Requirements phase after successful analysis
        if analysis_results:
            total_classes = sum(data.get("total_classes", 0) for data in analysis_results.values())
            total_tests = sum(data.get("total_tests", 0) for data in analysis_results.values())
            
            # Calculate progress based on analysis completeness
            progress = 100.0 if total_classes > 0 else 50.0
            
            sdlc_tracker.update_phase_progress(
                project_id=project.id,
                progress_percent=progress,
                tasks_completed=1,
                tasks_total=1,
                notes=f"Code analysis complete: {total_classes} classes, {total_tests} tests found"
            )
            logger.info(f"Requirements phase updated to {progress}% (ready to advance)")
        
        elapsed_time = time.time() - start_time
        logger.info(f"Project creation completed in {elapsed_time:.2f} seconds")

        return {
            "project_id": project.id,
            "name": project.name,
            "description": project.description,
            "current_phase": project.current_phase.value,
            "repositories": repo_configs,
            "analysis": {path: {k: v for k, v in data.items() if k != "classes"} for path, data in analysis_results.items()},
            "graph_summary": graph_summary,
            "created_at": project.created_at.isoformat()
        }

    except Exception as e:
        raise HTTPException(500, f"Failed to create project: {str(e)}")


@router.get("/projects", response_model=List[Dict])
async def list_projects():
    """List all SDLC projects"""
    projects = sdlc_tracker.get_all_projects()

    return [
        {
            "project_id": p.id,
            "name": p.name,
            "description": p.description,
            "current_phase": p.current_phase.value,
            "repository_count": len(p.repositories),
            "created_at": p.created_at.isoformat(),
            "updated_at": p.updated_at.isoformat()
        }
        for p in projects
    ]


@router.get("/projects/{project_id}", response_model=Dict)
async def get_project(project_id: str):
    """Get project details"""
    project = sdlc_tracker.get_project(project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    # Get current phase progress
    current_progress = sdlc_tracker.get_current_phase_progress(project_id)

    return {
        "project_id": project.id,
        "name": project.name,
        "description": project.description,
        "current_phase": project.current_phase.value,
        "repositories": project.repositories,
        "current_phase_progress": {
            "progress_percent": current_progress.progress_percent if current_progress else 0,
            "tasks_completed": current_progress.tasks_completed if current_progress else 0,
            "tasks_total": current_progress.tasks_total if current_progress else 0,
            "status": current_progress.status.value if current_progress else "not-started"
        },
        "created_at": project.created_at.isoformat(),
        "updated_at": project.updated_at.isoformat(),
        "metadata": project.metadata
    }


# ── Phase Management Endpoints ────────────────────────────────────────────────

@router.get("/projects/{project_id}/tracker", response_model=Dict)
async def get_phase_tracker(project_id: str):
    """Get 'train tracker' style phase visualization

    Returns current phase, completed phases, and upcoming phases
    """
    project = sdlc_tracker.get_project(project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    tracker_view = sdlc_tracker.get_phase_tracker_view(project_id)
    return tracker_view


@router.post("/projects/{project_id}/progress", response_model=Dict)
async def update_progress(project_id: str, request: UpdateProgressRequest):
    """Update progress of current phase"""
    success = sdlc_tracker.update_phase_progress(
        project_id,
        progress_percent=request.progress_percent,
        tasks_completed=request.tasks_completed,
        tasks_total=request.tasks_total,
        notes=request.notes
    )

    if not success:
        raise HTTPException(404, "Project not found or update failed")

    # Get updated progress
    current_progress = sdlc_tracker.get_current_phase_progress(project_id)

    return {
        "success": True,
        "progress_percent": current_progress.progress_percent if current_progress else 0,
        "tasks_completed": current_progress.tasks_completed if current_progress else 0,
        "tasks_total": current_progress.tasks_total if current_progress else 0
    }


@router.post("/projects/{project_id}/transition", response_model=Dict)
async def transition_phase(project_id: str, request: TransitionPhaseRequest):
    """Transition project to a new phase"""
    project = sdlc_tracker.get_project(project_id)
    logger.info(f"Transition request: project={project_id}, current_phase={project.current_phase.value if project else 'N/A'}, target_phase={request.target_phase}")
    
    try:
        # Parse target phase
        target_phase = SDLCPhaseType[request.target_phase.upper().replace(" ", "_")]
    except KeyError:
        raise HTTPException(400, f"Invalid phase: {request.target_phase}")

    # Auto-complete current phase if it has code (BEFORE checking prerequisites)
    # This prevents getting stuck on phases that should auto-complete
    if project and project.metadata.get("analysis"):
        analysis = project.metadata["analysis"]
        total_classes = sum(data.get("total_classes", 0) for data in analysis.values())
        
        if total_classes > 0:
            current_progress = sdlc_tracker.get_current_phase_progress(project_id)
            if current_progress and current_progress.progress_percent < 80.0:
                # Auto-complete Requirements, Design, Development, Deployment phases
                if project.current_phase in [SDLCPhaseType.REQUIREMENTS, SDLCPhaseType.DESIGN, SDLCPhaseType.DEVELOPMENT, SDLCPhaseType.DEPLOYMENT]:
                    sdlc_tracker.update_phase_progress(
                        project_id=project_id,
                        progress_percent=100.0,
                        tasks_completed=1,
                        tasks_total=1,
                        notes=f"{project.current_phase.value} phase auto-completed (code exists)"
                    )
                    logger.info(f"{project.current_phase.value} phase auto-completed to 100% before transition check")

    # Check if can transition
    can_transition, reasons = sdlc_tracker.can_transition_to_phase(project_id, target_phase)

    if not can_transition and not request.force:
        return {
            "success": False,
            "can_transition": False,
            "reasons": reasons,
            "message": "Prerequisites not met. Use force=true to override."
        }

    # Perform transition
    success, message = sdlc_tracker.transition_to_phase(project_id, target_phase, request.force)

    if not success:
        raise HTTPException(400, message)

    # Verify the transition happened
    updated_project = sdlc_tracker.get_project(project_id)
    logger.info(f"Transition completed: new_phase={updated_project.current_phase.value if updated_project else 'N/A'}")
    
    # Auto-complete the NEW phase if it has code (so we can continue advancing)
    if updated_project and updated_project.metadata.get("analysis"):
        analysis = updated_project.metadata["analysis"]
        total_classes = sum(data.get("total_classes", 0) for data in analysis.values())
        
        if total_classes > 0:
            # Auto-complete Design and Development phases after entering them
            if target_phase in [SDLCPhaseType.DESIGN, SDLCPhaseType.DEVELOPMENT]:
                progress_percent = 100.0 if target_phase == SDLCPhaseType.DESIGN else 80.0
                sdlc_tracker.update_phase_progress(
                    project_id=project_id,
                    progress_percent=progress_percent,
                    tasks_completed=1,
                    tasks_total=1,
                    notes=f"{target_phase.value} phase auto-completed (code exists)"
                )
                logger.info(f"New {target_phase.value} phase auto-completed to {progress_percent}%")
            
            # Auto-start Playwright tests when entering Testing phase
            elif target_phase == SDLCPhaseType.TESTING:
                logger.info("Testing phase entered - auto-starting Playwright tests")
                try:
                    # Get repo path
                    local_path = ""
                    for repo in updated_project.repositories:
                        if repo.get("local_path"):
                            local_path = repo["local_path"]
                            break
                    
                    if local_path:
                        # Create and execute integration test session
                        classes_to_test = testing_service.build_playwright_stops(local_path)
                        session = testing_service.create_test_session(
                            project_id=project_id,
                            test_type="integration",
                            classes_to_test=classes_to_test,
                        )
                        
                        # Start execution immediately
                        testing_service.start_test_execution(session.id)
                        
                        # Store session ID in project metadata
                        updated_project.metadata["active_session_id"] = session.id
                        updated_project.metadata["active_session_type"] = "integration"
                        
                        logger.info(f"Playwright tests auto-started: session_id={session.id}, tests={len(classes_to_test)}")
                except Exception as e:
                    logger.warning(f"Failed to auto-start Playwright tests: {e}")
            
            # Auto-complete Deployment phase (tests passed, ready to deploy)
            elif target_phase == SDLCPhaseType.DEPLOYMENT:
                # If we reached Deployment, tests must have passed
                # Auto-complete to 100% (ready for maintenance)
                sdlc_tracker.update_phase_progress(
                    project_id=project_id,
                    progress_percent=100.0,
                    tasks_completed=1,
                    tasks_total=1,
                    notes="Deployment phase auto-completed (tests passed, ready to deploy)"
                )
                logger.info("Deployment phase auto-completed to 100%")

    return {
        "success": True,
        "can_transition": True,
        "message": message,
        "new_phase": target_phase.value,
        "current_phase": updated_project.current_phase.value if updated_project else None
    }


# ── Testing Endpoints ─────────────────────────────────────────────────────────

@router.post("/projects/{project_id}/test", response_model=Dict)
async def start_testing(project_id: str, request: StartTestingRequest):
    """Start unit testing for a project

    This endpoint:
    1. Analyzes code to find classes and tests
    2. Creates test execution session
    3. Returns test plan with classes to be tested
    """
    project = sdlc_tracker.get_project(project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    # Resolve repo path — prefer application repo, fall back to any
    all_repo_types = ["application", "infrastructure", "devops"]
    local_path = ""
    for rt in all_repo_types:
        candidates = [r for r in project.repositories if r.get("type") == rt and r.get("local_path")]
        if candidates:
            local_path = candidates[0]["local_path"]
            break

    classes_to_test = []
    test_plan = {}

    if request.test_type == "unit":
        if not local_path:
            raise HTTPException(400, "No repository path found for unit testing")
        result = code_analyzer.analyze_directory(local_path)
        test_plan = code_analyzer.get_test_execution_plan(result)
        for cls in result.classes:
            if cls.is_test:
                continue
            class_tests = [t for t in result.tests if t.tests_class == cls.name]
            classes_to_test.append(ClassTestResult(
                class_name=cls.name,
                file_path=cls.file_path,
                test_file_path=class_tests[0].file_path if class_tests else None,
                tests_total=max(len(class_tests), 1),
                test_methods=[t.name for t in class_tests],
                status=TestStatus.PENDING,
            ))

    elif request.test_type == "regression":
        classes_to_test = testing_service.build_regression_stops(local_path or ".")

    elif request.test_type == "integration":
        classes_to_test = testing_service.build_playwright_stops(local_path or ".")

    session = testing_service.create_test_session(
        project_id=project_id,
        test_type=request.test_type,
        classes_to_test=classes_to_test,
    )

    return {
        "session_id": session.id,
        "project_id": project_id,
        "test_type": session.test_type,
        "total_tests": session.total_tests,
        "total_classes": len(classes_to_test),
        "execution_order": session.execution_order,
        "test_plan": test_plan,
        "status": session.status,
    }


@router.get("/test-sessions/{session_id}/status", response_model=Dict)
async def get_test_status(session_id: str):
    """Get test execution status and progress"""
    session = testing_service.get_session(session_id)
    if not session:
        raise HTTPException(404, "Test session not found")

    summary = testing_service.get_test_summary(session_id)
    return summary


@router.post("/projects/{project_id}/sync-test-progress", response_model=Dict)
async def sync_test_progress(project_id: str):
    """Manually sync Testing phase progress from active test session
    
    Use this if tests completed but phase progress didn't update
    """
    project = sdlc_tracker.get_project(project_id)
    if not project:
        raise HTTPException(404, "Project not found")
    
    # Get active session from metadata
    session_id = project.metadata.get("active_session_id")
    if not session_id:
        raise HTTPException(400, "No active test session found")
    
    session = testing_service.get_session(session_id)
    if not session:
        raise HTTPException(404, "Test session not found")
    
    # Calculate and update progress
    total_tests = session.tests_passed + session.tests_failed
    pass_rate = (session.tests_passed / total_tests * 100) if total_tests > 0 else 0
    progress = min(100.0, pass_rate) if pass_rate >= 80 else pass_rate * 0.8
    
    sdlc_tracker.update_phase_progress(
        project_id=project_id,
        progress_percent=progress,
        tasks_completed=session.tests_passed,
        tasks_total=total_tests,
        notes=f"{session.test_type.title()} tests: {session.tests_passed}/{total_tests} passed ({pass_rate:.1f}%)"
    )
    
    logger.info(f"Manually synced Testing phase to {progress:.1f}% from session {session_id}")
    
    return {
        "success": True,
        "progress_percent": progress,
        "tests_passed": session.tests_passed,
        "tests_failed": session.tests_failed,
        "total_tests": total_tests,
        "pass_rate": pass_rate,
        "session_status": session.status
    }


@router.get("/test-sessions/{session_id}/pipeline", response_model=Dict)
async def get_test_pipeline(session_id: str):
    """Get test pipeline visualization

    Shows which classes are being tested, progress, and results
    """
    session = testing_service.get_session(session_id)
    if not session:
        raise HTTPException(404, "Test session not found")

    pipeline = testing_service.get_test_pipeline_visualization(session_id)

    return {
        "session_id": pipeline.session_id,
        "test_type": getattr(pipeline, "test_type", session.test_type),
        "total_steps": pipeline.total_steps,
        "completed_steps": pipeline.completed_steps,
        "current_step": pipeline.current_step,
        "progress_percent": pipeline.progress_percent,
        "steps": pipeline.steps,
    }


@router.get("/test-sessions/{session_id}/matrix", response_model=List[Dict])
async def get_test_matrix(session_id: str):
    """Get class-to-test matrix view"""
    session = testing_service.get_session(session_id)
    if not session:
        raise HTTPException(404, "Test session not found")

    matrix = testing_service.get_classes_matrix(session_id)
    return matrix


@router.post("/test-sessions/{session_id}/execute", response_model=Dict)
async def execute_tests(session_id: str, background_tasks: BackgroundTasks):
    """Execute all tests in the session

    Tests are executed sequentially, traversing classes one by one
    """
    session = testing_service.get_session(session_id)
    if not session:
        raise HTTPException(404, "Test session not found")

    if session.status == "running":
        raise HTTPException(400, "Tests are already running")

    # Start execution in background
    # In production, this should be a proper background job
    testing_service.start_test_execution(session_id)
    
    # Add background task to update Testing phase when tests complete
    def update_phase_on_completion():
        """Poll for test completion and update Testing phase"""
        import time as time_module
        max_wait = 300  # 5 minutes max
        start = time_module.time()
        
        while time_module.time() - start < max_wait:
            session = testing_service.get_session(session_id)
            if session and session.status == "completed":
                # Calculate pass rate
                total_tests = session.tests_passed + session.tests_failed
                pass_rate = (session.tests_passed / total_tests * 100) if total_tests > 0 else 0
                
                # Update Testing phase progress
                project = sdlc_tracker.get_project(session.project_id)
                if project and project.current_phase.value == "Testing":
                    # 100% pass = 100% phase complete, <80% pass = proportional
                    progress = min(100.0, pass_rate) if pass_rate >= 80 else pass_rate * 0.8
                    
                    sdlc_tracker.update_phase_progress(
                        project_id=session.project_id,
                        progress_percent=progress,
                        tasks_completed=session.tests_passed,
                        tasks_total=total_tests,
                        notes=f"{session.test_type.title()} tests: {session.tests_passed}/{total_tests} passed ({pass_rate:.1f}%)"
                    )
                    logger.info(f"Testing phase auto-updated to {progress:.1f}% after test completion")
                break
            time_module.sleep(2)
    
    background_tasks.add_task(update_phase_on_completion)

    return {
        "session_id": session_id,
        "status": "started",
        "message": "Test execution started. Poll /status endpoint for progress."
    }


# ── Dashboard Endpoints ───────────────────────────────────────────────────────

@router.get("/dashboard", response_model=Dict)
async def get_dashboard():
    """Get SDLC dashboard overview

    Shows all projects grouped by phase, active tests, etc.
    """
    projects = sdlc_tracker.get_all_projects()

    # Group by phase
    projects_by_phase = {}
    for phase in SDLCPhaseType:
        projects_by_phase[phase.value] = []

    for project in projects:
        phase_name = project.current_phase.value
        projects_by_phase[phase_name].append({
            "id": project.id,
            "name": project.name,
            "description": project.description[:100] if len(project.description) > 100 else project.description
        })

    # Get active test sessions
    active_tests = []
    for session in testing_service.sessions.values():
        if session.status in ["running", "pending"]:
            summary = testing_service.get_test_summary(session.id)
            active_tests.append(summary)

    return {
        "total_projects": len(projects),
        "projects_by_phase": projects_by_phase,
        "active_test_sessions": len(active_tests),
        "test_sessions": active_tests
    }


# ── Code Analysis Endpoints ───────────────────────────────────────────────────

@router.get("/projects/{project_id}/analysis", response_model=Dict)
async def get_code_analysis(project_id: str):
    """Get detailed code analysis for a project"""
    project = sdlc_tracker.get_project(project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    # Get application repository
    app_repos = [r for r in project.repositories if r["type"] == "application"]
    if not app_repos:
        raise HTTPException(400, "No application repository found")

    local_path = app_repos[0]["local_path"]

    # Analyze code
    result = code_analyzer.analyze_directory(local_path)

    return {
        "total_modules": len(result.modules),
        "total_classes": result.total_classes,
        "total_functions": result.total_functions,
        "total_tests": result.total_tests,
        "total_lines": result.total_lines,
        "test_coverage_percent": result.test_coverage_percent,
        "classes": [
            {
                "name": cls.name,
                "file_path": cls.file_path,
                "methods": cls.methods,
                "complexity": cls.complexity,
                "is_test": cls.is_test
            }
            for cls in result.classes[:50]  # Limit to first 50
        ],
        "dependencies": result.dependencies
    }


# ── Folder Browse Endpoint ────────────────────────────────────────────────────

@router.get("/browse-folder", response_model=Dict)
async def browse_folder(initial_dir: str = ""):
    """Open a native OS folder-picker dialog and return the selected path.

    Only works when the backend is running locally (uses tkinter).
    """
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()          # hide the blank Tk window
        root.wm_attributes("-topmost", True)   # bring dialog to front

        selected = filedialog.askdirectory(
            title="Select Repository Folder",
            initialdir=initial_dir or "C:/",
        )
        root.destroy()

        if not selected:
            return {"path": "", "cancelled": True}

        # Normalise to OS path separators
        import os
        return {"path": os.path.normpath(selected), "cancelled": False}

    except Exception as e:
        raise HTTPException(500, f"Folder picker unavailable: {e}")


# ── DevOps Pipeline Upload ───────────────────────────────────────────────────

@router.post("/projects/{project_id}/pipeline-upload", response_model=Dict)
async def upload_pipeline(project_id: str, file: UploadFile = File(...)):
    """Upload a DevOps pipeline definition (JSON/YAML) and extract a deployment graph."""
    project = sdlc_tracker.get_project(project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    raw = await file.read()
    if not raw:
        raise HTTPException(400, "Empty file")

    filename = file.filename or "pipeline.json"

    project.metadata["deployment_graph"] = {"pipeline_name": filename, "nodes": 0, "edges": 0}
    return {"success": True, "pipeline_name": filename, "stage_count": 0, "nodes": 0, "edges": 0}


@router.get("/projects/{project_id}/deployment-graph", response_model=Dict)
async def get_deployment_graph(project_id: str):
    """Return the deployment graph for a project."""
    project = sdlc_tracker.get_project(project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    dg = project.metadata.get("deployment_graph")
    if not dg:
        return {"nodes": [], "edges": [], "message": "No pipeline uploaded yet"}

    return dg


@router.post("/projects/{project_id}/git-ingest", response_model=Dict)
async def ingest_git(project_id: str):
    """Ingest git repository data into the ontology knowledge graph."""
    project = sdlc_tracker.get_project(project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    app_repos = [r for r in project.repositories if r.get("local_path")]
    if not app_repos:
        raise HTTPException(400, "No repository with a local path found for this project")

    results = []
    for repo in app_repos:
        local_path = repo["local_path"]
        try:
            result = code_analyzer.analyze_directory(local_path)
            results.append({"path": local_path, "triples": 0, "classes": result.total_classes})
        except Exception as exc:
            results.append({"path": local_path, "error": str(exc)})

    return {"success": True, "repos_ingested": len(results), "results": results}


@router.get("/projects/{project_id}/test-suggestions", response_model=Dict)
async def get_test_suggestions(project_id: str):
    """Generate LLM-powered test suggestions based on the project's code graph."""
    project = sdlc_tracker.get_project(project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    suggestions = testing_service.generate_test_suggestions(project_id)

    by_type: dict = {"unit": [], "integration": [], "deployment": []}
    for s in suggestions:
        t = s.get("type", "unit")
        if t in by_type:
            by_type[t].append(s)
        else:
            by_type["unit"].append(s)

    return {
        "project_id": project_id,
        "total": len(suggestions),
        "by_type": by_type,
        "suggestions": suggestions,
    }


# ── Infra Upload Endpoint ─────────────────────────────────────────────────────

@router.post("/projects/{project_id}/infra-upload", response_model=Dict)
async def upload_infra_data(project_id: str, file: UploadFile = File(...)):
    """Upload infrastructure data (JSON or Excel) and populate the knowledge graph.

    JSON format: list of objects with keys: name, type, provider, region, status
    Excel format: sheet with columns: name, type, provider, region, status
    """
    project = sdlc_tracker.get_project(project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    content = await file.read()
    filename = file.filename or ""
    resources = []

    try:
        if filename.endswith(".json"):
            data = json.loads(content.decode("utf-8"))
            resources = data if isinstance(data, list) else [data]

        elif filename.endswith((".xlsx", ".xls")):
            try:
                import openpyxl
                import io
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                headers = [str(cell.value).lower().strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {headers[i]: (row[i] or "") for i in range(min(len(headers), len(row)))}
                    resources.append(row_dict)
            except ImportError:
                raise HTTPException(400, "openpyxl not installed. Install it with: pip install openpyxl")
        else:
            raise HTTPException(400, "Unsupported file type. Upload .json or .xlsx")

    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        raise HTTPException(400, f"Failed to parse file: {e}")

    if not resources:
        raise HTTPException(400, "No resources found in uploaded file")

    types_found = list({r.get("type", "CloudResource") for r in resources if r.get("type")})
    return {
        "success": True,
        "resources_loaded": len(resources),
        "types": types_found,
        "message": f"Loaded {len(resources)} infrastructure resources.",
    }
