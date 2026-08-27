"""File-based ontology loader — parse JSON, Excel (.xlsx), and CSV into Neo4j.

Expected JSON format:
  [
    {
      "label": "Service",
      "properties": {"name": "ClaimService", "status": "active"},
      "relationships": [
        {"type": "DEPENDS_ON", "targetExternalId": "repo:claims-api", "props": {}}
      ]
    }
  ]

Excel format:
  - One sheet per label (e.g. "Service", "Repository")
  - First column must be "externalId"
  - Optional "_Relationships" sheet with columns: fromExternalId, fromLabel, relType, toExternalId, toLabel

CSV format:
  - Single label per file (label inferred from filename without extension)
  - First row = column headers, first column = externalId
"""
from __future__ import annotations

import csv
import io
import json
import logging
import uuid
from typing import Any

log = logging.getLogger(__name__)


# ── Public API ────────────────────────────────────────────────────────────────

def load_file_bytes(raw: bytes, filename: str, version_id: str | None = None) -> dict[str, int]:
    """Dispatch to the correct parser based on file extension."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "json":
        records = _parse_json(raw)
    elif ext in ("xlsx", "xls"):
        records = _parse_excel(raw)
    elif ext == "csv":
        label = filename.rsplit(".", 1)[0].replace("-", "_").replace(" ", "_")
        records = _parse_csv(raw, label)
    else:
        raise ValueError(f"Unsupported file type: .{ext}. Supported: .json, .xlsx, .csv")
    return load_json_records(records, version_id)


def load_json_records(records: list[dict], version_id: str | None = None) -> dict[str, int]:
    """Write normalised records to Neo4j. Returns stats dict."""
    from src.graph import neo4j_client as neo4j

    nodes_added = nodes_updated = rels_added = 0
    for rec in records:
        label = rec.get("label", "Unknown")
        props: dict[str, Any] = rec.get("properties", {})
        ext_id = props.get("externalId") or f"file:{uuid.uuid4()}"

        existing = _node_exists(label, ext_id)
        neo4j.upsert_node_with_version(label, ext_id, props, version_id=version_id)
        if existing:
            nodes_updated += 1
        else:
            nodes_added += 1

        for rel in rec.get("relationships", []):
            try:
                neo4j.upsert_relationship(
                    from_label=label,
                    from_eid=ext_id,
                    to_label=rel.get("targetLabel", "Unknown"),
                    to_eid=rel["targetExternalId"],
                    rel_type=rel["type"],
                    props=rel.get("props"),
                    provenance={
                        "source": "file_upload",
                        "discoveredBy": "file_load_service",
                        "confidence": 1.0,
                        "factType": "known",
                    },
                )
                rels_added += 1
            except Exception as exc:
                log.warning("Relationship skip: %s", exc)

    total = _count_nodes()
    return {
        "nodesAdded": nodes_added,
        "nodesUpdated": nodes_updated,
        "relsAdded": rels_added,
        "totalNodes": total,
    }


def preview_file_bytes(raw: bytes, filename: str, rows: int = 5) -> dict:
    """Return a preview of parsed records without writing to Neo4j."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    try:
        if ext == "json":
            records = _parse_json(raw)
        elif ext in ("xlsx", "xls"):
            records = _parse_excel(raw)
        elif ext == "csv":
            label = filename.rsplit(".", 1)[0]
            records = _parse_csv(raw, label)
        else:
            raise ValueError(f"Unsupported: .{ext}")
        return {"total": len(records), "preview": records[:rows]}
    except Exception as exc:
        return {"error": str(exc), "total": 0, "preview": []}


# ── Parsers ───────────────────────────────────────────────────────────────────

def _parse_json(raw: bytes) -> list[dict]:
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid JSON: {exc}")
    if isinstance(data, dict):
        # Support {"nodes": [...]} envelope
        data = data.get("nodes", list(data.values())[0] if data else [])
    if not isinstance(data, list):
        raise ValueError("JSON must be an array of node records")
    return [_normalise_record(r) for r in data if isinstance(r, dict)]


def _parse_excel(raw: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise ValueError("openpyxl is not installed. Run: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    records: list[dict] = []
    rel_rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        headers_row = next(rows_iter, None)
        if not headers_row:
            continue
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(headers_row)]

        if sheet_name == "_Relationships":
            for row in rows_iter:
                d = dict(zip(headers, row))
                if any(d.values()):
                    rel_rows.append(d)
            continue

        label = sheet_name
        for row in rows_iter:
            d = dict(zip(headers, row))
            if not any(v is not None for v in d.values()):
                continue
            props = {k: v for k, v in d.items() if v is not None}
            if "externalId" not in props:
                props["externalId"] = f"excel:{uuid.uuid4()}"
            records.append({"label": label, "properties": props, "relationships": []})

    # Attach relationships to their source nodes
    node_map = {r["properties"].get("externalId"): r for r in records}
    for rel in rel_rows:
        src_eid = rel.get("fromExternalId")
        if src_eid and src_eid in node_map:
            node_map[src_eid]["relationships"].append({
                "type": rel.get("relType", "RELATED_TO"),
                "targetLabel": rel.get("toLabel", "Unknown"),
                "targetExternalId": rel.get("toExternalId", ""),
                "props": {},
            })

    return records


def _parse_csv(raw: bytes, label: str) -> list[dict]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    records = []
    for row in reader:
        props = {k.strip(): v for k, v in row.items() if v and v.strip()}
        if not props:
            continue
        if "externalId" not in props:
            props["externalId"] = f"csv:{uuid.uuid4()}"
        records.append({"label": label.capitalize(), "properties": props, "relationships": []})
    return records


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalise_record(raw: dict) -> dict:
    """Ensure a record has the expected structure."""
    label = raw.get("label") or raw.get("type") or raw.get("nodeType") or "Unknown"
    props = raw.get("properties") or raw.get("props") or {}
    # Handle flat records where properties are at the top level
    if not props:
        props = {k: v for k, v in raw.items() if k not in ("label", "type", "nodeType", "relationships")}
    if "externalId" not in props:
        props["externalId"] = f"file:{uuid.uuid4()}"
    rels = raw.get("relationships", [])
    return {"label": label, "properties": props, "relationships": rels}


def _node_exists(label: str, ext_id: str) -> bool:
    try:
        from src.graph import neo4j_client as neo4j
        cypher = f"MATCH (n:{label} {{externalId: $eid}}) RETURN 1 LIMIT 1"
        with neo4j.session() as s:
            return s.run(cypher, eid=ext_id).single() is not None
    except Exception:
        return False


def _count_nodes() -> int:
    try:
        from src.graph import neo4j_client as neo4j
        with neo4j.session() as s:
            result = s.run("MATCH (n) RETURN count(n) AS c").single()
            return int(result["c"]) if result else 0
    except Exception:
        return 0
